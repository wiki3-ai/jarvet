#!/usr/bin/env python3
"""Build a local school-program index from IPEDS and the O*NET CIP crosswalk.

Downloads the IPEDS institutional directory (HD) and completions (C_A) files
plus the official O*NET Education CIP-to-SOC crosswalk, then joins them into a
compact SQLite index: institutions with coordinates, and program rows keyed by
O*NET-SOC code with recent-award counts. This replaces scraping My Next Move,
whose local-training table is itself derived from IPEDS completions and the
same crosswalk.
"""
from __future__ import annotations

import csv
import io
import re
import sqlite3
import sys
import urllib.request
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CACHE = ROOT / ".cache"
DATABASE = CACHE / "ipeds.sqlite"
MARKER = CACHE / "ipeds.ready"
YEAR = 2024
HD_URL = f"https://nces.ed.gov/ipeds/datacenter/data/HD{YEAR}.zip"
COMPLETIONS_URL = f"https://nces.ed.gov/ipeds/datacenter/data/C{YEAR}_A.zip"
CROSSWALK_URL = (
    "https://www.onetcenter.org/crosswalks/cip/Education_CIP_to_ONET_SOC.xlsx"
)
INDEX_VERSION = 1


def download(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "jarvet/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def column_index(reference: str) -> int:
    index = 0
    for character in re.match(r"[A-Z]+", reference).group():
        index = index * 26 + ord(character) - 64
    return index - 1


def crosswalk_rows(workbook: bytes) -> list[tuple[str, str, str]]:
    from openpyxl import load_workbook

    sheet = load_workbook(io.BytesIO(workbook), read_only=True).active
    triples: list[tuple[str, str, str]] = []
    for row in sheet.iter_rows(values_only=True):
        cip, cip_title, soc = row[0], row[1], row[2]
        if isinstance(cip, str) and isinstance(soc, str) and re.match(
            r"^\d{2}\.\d{4}$", cip.strip()
        ) and re.match(r"^\d{2}-\d{4}\.\d{2}$", soc.strip()):
            triples.append((cip.strip(), (cip_title or "").strip(), soc.strip()))
    return triples


def main() -> None:
    marker = f"{INDEX_VERSION}:{YEAR}"
    if DATABASE.exists() and MARKER.exists() and MARKER.read_text() == marker:
        print("IPEDS index is ready.")
        return

    CACHE.mkdir(parents=True, exist_ok=True)
    print("Downloading IPEDS directory...")
    hd_zip = zipfile.ZipFile(io.BytesIO(download(HD_URL)))
    hd_name = next(n for n in hd_zip.namelist() if n.lower().endswith(".csv"))
    hd_rows = list(csv.DictReader(
        (line.decode("utf-8-sig", errors="replace") for line in hd_zip.open(hd_name))
    ))

    print("Downloading IPEDS completions...")
    completions_zip = zipfile.ZipFile(io.BytesIO(download(COMPLETIONS_URL)))
    completions_name = next(
        n for n in completions_zip.namelist() if n.lower().endswith(".csv")
    )
    completions_rows = list(csv.DictReader(
        (line.decode("utf-8-sig", errors="replace") for line in completions_zip.open(completions_name))
    ))

    print("Downloading O*NET CIP-to-SOC crosswalk...")
    crosswalk = crosswalk_rows(download(CROSSWALK_URL))
    cip_to_soc: dict[str, set[str]] = {}
    cip_titles: dict[str, str] = {}
    for cip, cip_title, soc in crosswalk:
        cip_to_soc.setdefault(cip, set()).add(soc)
        cip_titles.setdefault(cip, cip_title)

    connection = sqlite3.connect(DATABASE)
    connection.execute(
        "CREATE TABLE institutions ("
        "unitid TEXT PRIMARY KEY, name TEXT NOT NULL, city TEXT, state TEXT, "
        "zip TEXT, latitude REAL, longitude REAL, website TEXT, "
        "iclevel INTEGER, control INTEGER)"
    )
    connection.execute(
        "CREATE TABLE programs ("
        "unitid TEXT NOT NULL, cip TEXT NOT NULL, awlevel INTEGER NOT NULL, "
        "awards INTEGER NOT NULL, soc_codes TEXT NOT NULL, "
        "PRIMARY KEY (unitid, cip, awlevel))"
    )
    connection.execute(
        "CREATE INDEX programs_soc ON programs(soc_codes)"
    )
    connection.execute(
        "CREATE TABLE soc_index (soc TEXT PRIMARY KEY, program_count INTEGER)"
    )
    connection.execute(
        "CREATE TABLE cip_titles (cip TEXT PRIMARY KEY, title TEXT NOT NULL)"
    )

    def first(row: dict, *names: str) -> str:
        for name in names:
            value = row.get(name)
            if value is not None:
                return (value or "").strip()
        return ""

    institutions = 0
    for row in hd_rows:
        unitid = first(row, "UNITID")
        name = first(row, "INSTNM")
        if not unitid or not name:
            continue
        latitude = first(row, "LATITUDE")
        longitude = first(row, "LONGITUD")
        connection.execute(
            "INSERT OR REPLACE INTO institutions VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                unitid, name, first(row, "CITY"), first(row, "STABBR"),
                first(row, "ZIP")[:5],
                float(latitude) if latitude else None,
                float(longitude) if longitude else None,
                first(row, "WEBADDR"), int(first(row, "ICLEVEL") or 0),
                int(first(row, "CONTROL") or 0),
            ),
        )
        institutions += 1

    programs = 0
    soc_counts: dict[str, int] = {}
    for row in completions_rows:
        unitid = first(row, "UNITID")
        cip = first(row, "CIPCODE")
        awards = first(row, "CTOTALT")
        if not unitid or not cip or awards in ("", None):
            continue
        try:
            award_count = int(awards)
        except ValueError:
            continue
        if award_count <= 0:
            continue
        awlevel = int(first(row, "AWLEVEL") or 0)
        socs = sorted(cip_to_soc.get(cip, ()))
        if not socs:
            continue
        connection.execute(
            "INSERT OR REPLACE INTO programs VALUES (?,?,?,?,?)",
            (unitid, cip, awlevel, award_count, ",".join(socs)),
        )
        programs += 1
        for soc in socs:
            soc_counts[soc] = soc_counts.get(soc, 0) + 1

    connection.executemany(
        "INSERT OR REPLACE INTO soc_index VALUES (?,?)", sorted(soc_counts.items())
    )
    connection.executemany(
        "INSERT OR REPLACE INTO cip_titles VALUES (?,?)", sorted(cip_titles.items())
    )
    connection.commit()
    connection.close()
    MARKER.write_text(marker)
    print(
        f"Indexed {institutions:,} institutions and {programs:,} program rows "
        f"covering {len(soc_counts):,} O*NET-SOC codes."
    )


if __name__ == "__main__":
    sys.exit(main())
